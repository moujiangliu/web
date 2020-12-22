'''
读取excel的方式
数据类型：列表套列表
openpyxl：这个第三方库只能读取后缀为.xlsx的excel文件
'''

# 导包
import openpyxl

from tool.test import get_path


class ReadExcel():
    def get_excel(self):
        # wb --> 工作簿
        wb = openpyxl.load_workbook(get_path('data','data.xlsx'))
        print(wb)
        # ws --> sheet页
        ws = wb['测试用例']

        # 数据类型：列表套列表
        all_cases = []
        for i in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            # print(i)
            # 列表推导式
            '''
            [cell.value                         for cell in i]
            [对循环遍历的cell进行处理               循环遍历i值，得到每一个cell]
            然后把每一个cell.value追加到一个列表里面
            '''
            # 原生写法
            # row_list = []
            # for cell in i:
            #     row_list.append(cell.value)
            row_list = [cell.value for cell in i]
            all_cases.append(row_list)

        return all_cases


if __name__ == '__main__':
    readexcel = ReadExcel().get_excel()
    print(readexcel)


