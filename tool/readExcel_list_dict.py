'''
读取excel的方式
数据类型：列表套字典
'''

# 导包
import openpyxl
import os

from tool.test import get_path


class ReadExcel():
    def get_excel(self, head):
        # wb --> 工作簿
        # wb = openpyxl.load_workbook('../data/data.xlsx')

        # 加r''，\\，/；以下路径表示方法是绝对路径的，缺点：灵活性差；框架：架构一样的，相同的
        # 绝对路径：base_dir        /data/+'data.xlsx'
        # D:\Learn\Code\webautoTest\lesson6_2\ ==> 如果获取这个路径？？？
        # base_dir = os.path.dirname(os.path.dirname(__file__))
        # wb = openpyxl.load_workbook(base_dir + '/data' + '/data.xlsx')
        # os.sep：自动去匹配你的系统的斜线，/或者\，（自动区分mac还是windows）
        # wb = openpyxl.load_workbook(base_dir + '%sdata%s' %(os.sep, os.sep ) + '/data.xlsx')

        wb = openpyxl.load_workbook(get_path('data', 'data.xlsx'))

        # print(wb)
        # ws --> sheet页
        ws = wb['测试用例']

        '''
        行数怎么确定？--> 最大行，从第二行开始，到最大行 for i in range()
        列数怎么确定？--> head里面的元素个数决定有几列
        第一个字典：username:第2行，第1列的cell值；password：第2行，第2列的cell值
        第二个字典：username:第3行，第1列的cell值；password：第3行，第2列的cell值
        第三个字典：username:第4行，第1列的cell值；password：第4行，第2列的cell值
        '''
        all_cases = []
        # 控制行数
        for row in range(2, ws.max_row+1):
            data = {}
            # 控制列数的，head里面有多个元素，那么就有几列
            col = 1
            for key in head:
                # {'username':第二行第一列的值}
                data[key] = ws.cell(row,col).value
                col += 1
            all_cases.append(data)
        return all_cases




if __name__ == '__main__':
    readexcel = ReadExcel().get_excel(['username', 'password'])
    print(readexcel)




























